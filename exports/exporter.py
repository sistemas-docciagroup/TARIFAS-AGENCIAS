from pathlib import Path
from datetime import datetime
import pandas as pd


_EXPORT_DIR = Path(__file__).resolve().parent.parent / "data" / "exports"
_EXPORT_DIR.mkdir(parents=True, exist_ok=True)

COLUMNS_DISPLAY = [
    "agencia", "factura", "fecha_envio", "expedicion_agencia",
    "albaran_doccia", "pedido_doccia", "destino", "destinatario",
    "bultos", "kilos", "peso_facturable",
    "portes", "combustible", "seguro", "reexpedicion", "otros",
    "total_facturado", "estado_cruce", "observaciones",
]

COLUMN_LABELS = {
    "agencia": "Agencia",
    "factura": "Factura",
    "fecha_envio": "Fecha Envío",
    "expedicion_agencia": "Expedición Agencia",
    "albaran_doccia": "Albarán Doccia",
    "pedido_doccia": "Pedido Doccia",
    "destino": "Destino",
    "destinatario": "Destinatario",
    "bultos": "Bultos",
    "kilos": "Kilos",
    "peso_facturable": "Peso Fact.",
    "portes": "Portes €",
    "combustible": "Combustible €",
    "seguro": "Seguro €",
    "reexpedicion": "Reexpedición €",
    "otros": "Otros €",
    "total_facturado": "Total Facturado €",
    "estado_cruce": "Estado",
    "observaciones": "Observaciones",
}


def _build_df(records: list[dict]) -> pd.DataFrame:
    if not records:
        return pd.DataFrame(columns=COLUMNS_DISPLAY)
    df = pd.DataFrame(records)
    for col in COLUMNS_DISPLAY:
        if col not in df.columns:
            df[col] = None
    return df[COLUMNS_DISPLAY].rename(columns=COLUMN_LABELS)


def export_excel(records: list[dict], filename: str | None = None) -> Path:
    df = _build_df(records)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = _EXPORT_DIR / (filename or f"exportacion_{ts}.xlsx")

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Albaranes")
        ws = writer.sheets["Albaranes"]

        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        for col_idx, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[cell.column_letter].width = max(15, len(str(cell.value or "")) + 4)

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
            if row_idx % 2 == 0:
                for cell in row:
                    cell.fill = alt_fill

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    return out


def export_csv(records: list[dict], filename: str | None = None) -> Path:
    df = _build_df(records)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = _EXPORT_DIR / (filename or f"exportacion_{ts}.csv")
    df.to_csv(out, index=False, sep=";", encoding="utf-8-sig")
    return out
