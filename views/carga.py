import streamlit as st
import tempfile
import pandas as pd
from pathlib import Path

from parsers.detector import detect_agency, extract_text_from_file
from parsers import PARSERS
from database.db import SessionLocal
from database.models import Albaran

AGENCIAS = ["Molartrans", "DHL Parcel", "DHL Freight", "CEVA", "TDN", "DSV"]


def _run_ai_verification(pdf_path: Path, rows: list[dict], agency: str, key: str):
    """Ejecuta la verificación IA y guarda resultado en session_state."""
    from services.ai_verifier import verify_invoice
    with st.spinner("Verificando con IA el PDF…"):
        result = verify_invoice(pdf_path, rows, agency)
    st.session_state[key] = result


def _card(label, value, sub=None, bg="#fff", border="#e8ebf3", val_color="#1a1f36"):
    sub_html = f"<div style='font-size:0.78rem;color:#6b7280;margin-top:3px'>{sub}</div>" if sub else ""
    return (
        f"<div style='background:{bg};border:1px solid {border};border-radius:10px;"
        f"padding:14px 18px;min-width:0'>"
        f"<div style='font-size:0.70rem;font-weight:600;color:#9ca3af;text-transform:uppercase;"
        f"letter-spacing:.4px;margin-bottom:6px'>{label}</div>"
        f"<div style='font-size:1.15rem;font-weight:700;color:{val_color};word-break:break-word'>{value}</div>"
        f"{sub_html}</div>"
    )


def _show_ai_result(result: dict, rows: list[dict]):
    """Muestra el análisis detallado de verificación IA."""
    if result.get("error"):
        st.warning(f"⚠️ Verificación IA no disponible: {result['error']}")
        return

    fac_ext    = rows[0].get("factura", "") if rows else ""
    fac_pdf    = result.get("numero_factura", "—")
    base_imp   = result.get("base_imponible_pdf")
    iva_pdf    = result.get("iva_pdf")
    total_iva  = result.get("total_pdf_con_iva")
    total_ext  = result.get("total_extraido") or sum(r.get("total_facturado") or 0 for r in rows)
    dif        = result.get("diferencia", 0) or 0
    ok_num     = result.get("coincide_numero", False)
    ok_tot     = result.get("coincide_base", False)   # comparamos contra base imponible
    causas     = result.get("causas_desviacion") or []
    desglose   = result.get("desglose_pdf") or {}
    recom      = result.get("recomendacion", "")
    confianza  = result.get("confianza", "—")
    n_pdf      = result.get("lineas_pdf")
    n_ext      = result.get("lineas_extraidas") or len(rows)

    todo_ok = ok_num and ok_tot

    # Corrección automática del número de factura
    if not ok_num and fac_pdf and fac_pdf not in ("—", "", None):
        for r in rows:
            r["factura"] = fac_pdf

    bg_head  = "#f0fdf4" if todo_ok else "#fef2f2"
    bd_head  = "#86efac" if todo_ok else "#fca5a5"
    icon     = "✅" if todo_ok else "⚠️"
    titulo   = "Verificación IA — Extracción correcta" if todo_ok else "Verificación IA — Desviación detectada"

    # ── Cabecera ──────────────────────────────────────────────────────────────
    st.markdown(
        f"<div style='background:{bg_head};border:1px solid {bd_head};border-radius:10px;"
        f"padding:12px 18px;margin-bottom:12px;font-weight:700;font-size:0.95rem'>"
        f"{icon} {titulo} "
        f"<span style='font-size:0.72rem;color:#6b7280;font-weight:400'>· Confianza IA: {confianza}</span>"
        f"</div>",
        unsafe_allow_html=True,
    )

    # ── Fila 1: comparativa numérica ─────────────────────────────────────────
    num_sub       = ("✓ coincide" if ok_num
                     else f"extraído: <b>{fac_ext or 'no detectado'}</b> → corregido automáticamente")
    dif_color     = "#16a34a" if abs(dif or 0) < 0.10 else "#dc2626"
    dif_bg        = "#f0fdf4" if abs(dif or 0) < 0.10 else "#fef2f2"
    dif_border    = "#86efac" if abs(dif or 0) < 0.10 else "#fca5a5"

    c1, c2, c3, c4, c5, c6, c7 = st.columns(7)
    c1.markdown(_card("Nº Factura PDF", fac_pdf,
                      sub=num_sub,
                      bg="#f0fdf4" if ok_num else "#fef2f2",
                      border="#86efac" if ok_num else "#fca5a5"),
                unsafe_allow_html=True)
    c2.markdown(_card("Base imponible PDF",
                      f"{base_imp:,.2f} €" if base_imp is not None else "—",
                      sub="sin IVA"),
                unsafe_allow_html=True)
    c3.markdown(_card("IVA PDF",
                      f"{iva_pdf:,.2f} €" if iva_pdf is not None else "—"),
                unsafe_allow_html=True)
    c4.markdown(_card("Total PDF con IVA",
                      f"{total_iva:,.2f} €" if total_iva is not None else "—"),
                unsafe_allow_html=True)
    c5.markdown(_card("Total extraído", f"{total_ext:,.2f} €",
                      sub="sin IVA"),
                unsafe_allow_html=True)
    c6.markdown(_card("Diferencia (base − extraído)",
                      f"{dif:+,.2f} €",
                      sub="✓ cuadra" if abs(dif or 0) < 0.10 else "revisar extracción",
                      bg=dif_bg, border=dif_border, val_color=dif_color),
                unsafe_allow_html=True)
    c7.markdown(_card("Líneas extraídas",
                      f"{n_ext:,}",
                      sub=(f"PDF: {n_pdf:,} {'✓' if n_ext == n_pdf else '⚠️'}"
                           if n_pdf is not None else "—"),
                      bg="#f0fdf4" if n_ext == n_pdf else "#fff7ed",
                      border="#86efac" if n_ext == n_pdf else "#fdba74"),
                unsafe_allow_html=True)

    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

    # ── Desglose por concepto ─────────────────────────────────────────────────
    conceptos = {k: v for k, v in desglose.items() if v not in (None, 0, 0.0)}
    if conceptos:
        etiquetas = {
            "portes": "Portes", "combustible": "Combustible",
            "seguro": "Seguro", "otros": "Otros",
            "base_imponible": "Base imponible", "iva": "IVA",
        }
        items_html = "".join(
            f"<div style='flex:1;min-width:120px;background:#f8f9fc;border:1px solid #e8ebf3;"
            f"border-radius:8px;padding:10px 14px'>"
            f"<div style='font-size:0.68rem;font-weight:600;color:#9ca3af;text-transform:uppercase;"
            f"letter-spacing:.4px;margin-bottom:4px'>{etiquetas.get(k, k)}</div>"
            f"<div style='font-size:1.05rem;font-weight:700;color:#1a1f36'>{v:,.2f} €</div>"
            f"</div>"
            for k, v in conceptos.items()
        )
        st.markdown(
            f"<div style='margin-bottom:10px'>"
            f"<div style='font-size:0.8rem;font-weight:600;color:#374151;margin-bottom:8px'>"
            f"Desglose del PDF por concepto</div>"
            f"<div style='display:flex;gap:10px;flex-wrap:wrap'>{items_html}</div></div>",
            unsafe_allow_html=True,
        )

    # ── Causas de la desviación ───────────────────────────────────────────────
    if causas:
        causas_html = "".join(
            f"<li style='margin-bottom:4px'>{c}</li>" for c in causas
        )
        st.markdown(
            f"<div style='background:#fff7ed;border:1px solid #fdba74;border-radius:8px;"
            f"padding:12px 16px;margin-bottom:10px'>"
            f"<div style='font-size:0.8rem;font-weight:600;color:#92400e;margin-bottom:6px'>"
            f"Causas de la desviación identificadas</div>"
            f"<ul style='margin:0;padding-left:18px;font-size:0.85rem;color:#374151'>"
            f"{causas_html}</ul></div>",
            unsafe_allow_html=True,
        )

    # ── Recomendación ─────────────────────────────────────────────────────────
    if recom:
        st.markdown(
            f"<div style='background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;"
            f"padding:10px 16px;font-size:0.85rem;color:#1e40af'>"
            f"<b>Recomendación:</b> {recom}</div>",
            unsafe_allow_html=True,
        )

    # ── Descarga del informe de comparación ──────────────────────────────────
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
    excel_bytes = _build_report_excel(rows, result)
    fac = rows[0].get("factura", "factura") if rows else "factura"
    st.download_button(
        label="📥 Descargar informe de comparación (.xlsx)",
        data=excel_bytes,
        file_name=f"informe_verificacion_{fac}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=False,
    )


def _build_report_excel(rows: list[dict], ai_result: dict) -> bytes:
    """Genera un Excel con dos hojas: detalle de líneas + resumen IA."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Hoja 1: Detalle de líneas extraídas ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "Líneas extraídas"

    hdr_fill  = PatternFill("solid", fgColor="1A1F36")
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    num_align = Alignment(horizontal="right")
    thin      = Border(
        bottom=Side(style="thin", color="D1D5E8"),
        right=Side(style="thin", color="D1D5E8"),
    )

    headers = ["Nº", "Expedición agencia", "Albarán Doccia", "Fecha envío",
               "Destinatario", "Destino", "Bultos",
               "Portes €", "RC Combustible €", "Seguro €", "Contrareembolso €", "TOTAL €",
               "Estado cruce"]
    widths  = [5, 18, 14, 12, 28, 20, 7, 11, 15, 11, 16, 11, 16]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        ws1.column_dimensions[get_column_letter(col)].width = w

    total_sum = 0.0
    for i, r in enumerate(rows, 2):
        total = r.get("total_facturado") or 0
        total_sum += total
        vals = [
            i - 1,
            r.get("expedicion_agencia") or "",
            r.get("albaran_doccia") or "",
            r.get("fecha_envio") or "",
            r.get("destinatario") or "",
            r.get("destino") or "",
            r.get("bultos") or "",
            r.get("portes") or 0,
            r.get("combustible") or 0,   # RC recargo combustible
            r.get("seguro") or 0,
            r.get("otros") or 0,         # comisión de contrareembolso
            total,
            r.get("estado_cruce") or "",
        ]
        for col, v in enumerate(vals, 1):
            cell = ws1.cell(row=i, column=col, value=v)
            cell.border = thin
            if isinstance(v, float):
                cell.number_format = "#,##0.00"
                cell.alignment     = num_align
        # Fila alternada
        if i % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws1.cell(row=i, column=col).fill = PatternFill("solid", fgColor="F8F9FC")

    # Fila de totales
    tot_row = len(rows) + 2
    ws1.cell(row=tot_row, column=11, value="TOTAL EXTRAÍDO →").font = Font(bold=True)
    tc = ws1.cell(row=tot_row, column=12, value=total_sum)
    tc.font          = Font(bold=True)
    tc.number_format = "#,##0.00"
    tc.fill          = PatternFill("solid", fgColor="FEF9C3")
    tc.alignment     = num_align

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ── Hoja 2: Resumen verificación IA ──────────────────────────────────────
    ws2 = wb.create_sheet("Verificación IA")

    def _write(row, col, label, value, bold_val=False):
        lc = ws2.cell(row=row, column=col, value=label)
        lc.font      = Font(bold=True, size=10, color="374151")
        lc.alignment = Alignment(horizontal="left")
        vc = ws2.cell(row=row, column=col + 1, value=value)
        vc.font      = Font(bold=bold_val, size=10)
        vc.alignment = Alignment(horizontal="left")
        return vc

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 30

    title_cell = ws2.cell(row=1, column=1, value="INFORME DE VERIFICACIÓN IA")
    title_cell.font = Font(bold=True, size=13, color="1A1F36")
    ws2.merge_cells("A1:B1")

    base_imp  = ai_result.get("base_imponible_pdf")
    iva_pdf   = ai_result.get("iva_pdf")
    total_iva = ai_result.get("total_pdf_con_iva")
    dif       = ai_result.get("diferencia", 0) or 0

    datos = [
        ("Nº Factura PDF",          ai_result.get("numero_factura", "—")),
        ("Base imponible PDF",       f"{base_imp:,.2f} €" if base_imp else "—"),
        ("IVA PDF",                  f"{iva_pdf:,.2f} €" if iva_pdf else "—"),
        ("Total PDF con IVA",        f"{total_iva:,.2f} €" if total_iva else "—"),
        ("Total extraído (sin IVA)", f"{total_sum:,.2f} €"),
        ("Diferencia",               f"{dif:+,.2f} €"),
        ("Líneas en PDF",            ai_result.get("lineas_pdf", "—")),
        ("Líneas extraídas",         len(rows)),
        ("Coincide número factura",  "Sí" if ai_result.get("coincide_numero") else "No"),
        ("Base imponible cuadra",    "Sí" if ai_result.get("coincide_base") else "No"),
        ("Confianza IA",             ai_result.get("confianza", "—")),
    ]
    for r_i, (lbl, val) in enumerate(datos, 3):
        _write(r_i, 1, lbl, val)

    # Causas
    causas = ai_result.get("causas_desviacion") or []
    if causas:
        r_i = len(datos) + 4
        ws2.cell(row=r_i, column=1, value="CAUSAS DE LA DESVIACIÓN").font = Font(bold=True, size=11)
        for j, c in enumerate(causas, r_i + 1):
            ws2.cell(row=j, column=1, value=f"• {c}").font = Font(size=10)
            ws2.merge_cells(f"A{j}:B{j}")
        r_i = j + 1
    else:
        r_i = len(datos) + 4

    # Recomendación
    recom = ai_result.get("recomendacion", "")
    if recom:
        ws2.cell(row=r_i + 1, column=1, value="RECOMENDACIÓN").font = Font(bold=True, size=11)
        rc = ws2.cell(row=r_i + 2, column=1, value=recom)
        rc.font      = Font(size=10, italic=True)
        rc.alignment = Alignment(wrap_text=True)
        ws2.merge_cells(f"A{r_i + 2}:B{r_i + 2}")
        ws2.row_dimensions[r_i + 2].height = 40

    # Desglose
    desglose = {k: v for k, v in (ai_result.get("desglose_pdf") or {}).items()
                if v not in (None, 0, 0.0)}
    if desglose:
        r_i += 4
        ws2.cell(row=r_i, column=1, value="DESGLOSE PDF POR CONCEPTO").font = Font(bold=True, size=11)
        etq = {"portes": "Portes", "combustible": "Combustible",
               "seguro": "Seguro", "otros": "Otros"}
        for j, (k, v) in enumerate(desglose.items(), r_i + 1):
            ws2.cell(row=j, column=1, value=etq.get(k, k)).font = Font(bold=True, size=10)
            vc = ws2.cell(row=j, column=2, value=v)
            vc.number_format = "#,##0.00"
            vc.alignment     = num_align

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def render():
    st.title("Carga de Facturas")

    if "carga_upload_key" not in st.session_state:
        st.session_state["carga_upload_key"] = 0
    if "previews" not in st.session_state:
        st.session_state["previews"] = {}

    if st.session_state.get("carga_guardada"):
        st.success(st.session_state.pop("carga_guardada"))

    # ── Zona de subida ────────────────────────────────────────────────────────
    col_up, col_ag = st.columns([3, 1])
    with col_up:
        uploaded_files = st.file_uploader(
            "Facturas",
            type=["pdf", "xlsx", "xls", "csv"],
            accept_multiple_files=True,
            label_visibility="collapsed",
            key=f"carga_uploader_{st.session_state['carga_upload_key']}",
        )
    with col_ag:
        st.caption("Forzar agencia")
        agencia_forzada = st.selectbox(
            "Agencia", ["Detectar automáticamente"] + AGENCIAS,
            label_visibility="collapsed", key="carga_ag_forzada",
        )
        agencia_override = None if agencia_forzada == "Detectar automáticamente" else agencia_forzada

    if not uploaded_files and not st.session_state["previews"]:
        return

    # ── Procesar archivos nuevos ──────────────────────────────────────────────
    if uploaded_files:
        new_files = [f for f in uploaded_files if f.name not in st.session_state["previews"]]
        if new_files:
            bar = st.progress(0, text="Procesando…")
            for i, f in enumerate(new_files):
                bar.progress((i + 1) / len(new_files), text=f"Procesando {f.name}…")
                with tempfile.NamedTemporaryFile(delete=False, suffix=Path(f.name).suffix) as tmp:
                    tmp.write(f.read())
                    tmp_path = Path(tmp.name)
                try:
                    text   = extract_text_from_file(tmp_path)
                    agency = agencia_override or detect_agency(text, f.name)
                    if agency and agency in PARSERS:
                        rows = PARSERS[agency]().parse(tmp_path)
                        st.session_state["previews"][f.name] = {
                            "agency": agency, "rows": rows,
                            "path": str(tmp_path), "error": None,
                        }
                    else:
                        st.session_state["previews"][f.name] = {
                            "agency": None, "rows": [],
                            "path": str(tmp_path), "error": "Agencia no detectada",
                        }
                except Exception as e:
                    st.session_state["previews"][f.name] = {
                        "agency": None, "rows": [],
                        "path": str(tmp_path), "error": str(e),
                    }
            bar.empty()

    st.markdown("---")

    # ── Vista previa por archivo ──────────────────────────────────────────────
    for fname in list(st.session_state["previews"].keys()):
        data   = st.session_state["previews"][fname]
        agency = data["agency"]
        rows   = data["rows"]
        error  = data["error"]
        is_pdf = fname.lower().endswith(".pdf")
        ai_key = f"ai_verify_{fname}"

        # Título del expander
        if error and not rows:
            titulo = f"❌ **{fname}** — {error}"
        else:
            ok  = sum(1 for r in rows if r.get("estado_cruce") in ("ALBARAN_OK", "PEDIDO_OK", "EXPEDICION_AGENCIA"))
            err = sum(1 for r in rows if r.get("estado_cruce") in ("ERROR_LECTURA", "SIN_REFERENCIA", "DUDOSO"))
            ai_done  = ai_key in st.session_state
            ai_ok    = ai_done and st.session_state[ai_key].get("coincide_numero") and st.session_state[ai_key].get("coincide_total")
            ai_badge = " · ✅ IA verificado" if (ai_done and ai_ok) else (" · ⚠️ IA: discrepancia" if ai_done else "")
            titulo   = f"{'✅' if not err else '⚠️'} **{fname}** — {agency} · {len(rows):,} líneas · {ok:,} OK · {err:,} incidencias{ai_badge}"

        with st.expander(titulo, expanded=True):

            # ── Error sin filas ───────────────────────────────────────────────
            if error and not rows:
                st.error(error)
                ag_m = st.selectbox("Asignar agencia", AGENCIAS, key=f"ag_err_{fname}")
                c1, c2 = st.columns(2)
                with c1:
                    if st.button("Reintentar", key=f"retry_{fname}", use_container_width=True):
                        try:
                            new_rows = PARSERS[ag_m]().parse(Path(data["path"]))
                            st.session_state["previews"][fname] = {
                                "agency": ag_m, "rows": new_rows,
                                "path": data["path"], "error": None,
                            }
                            st.session_state.pop(ai_key, None)
                            st.rerun()
                        except Exception as e2:
                            st.error(str(e2))
                with c2:
                    if st.button("🗑️ Descartar", key=f"del_err_{fname}", use_container_width=True):
                        del st.session_state["previews"][fname]
                        st.rerun()
                continue

            # ── Agencia no detectada ──────────────────────────────────────────
            if not agency:
                st.warning("Agencia no detectada.")
                ag_m = st.selectbox("Asignar agencia", AGENCIAS, key=f"ag_nd_{fname}")
                c1, c2 = st.columns(2)
                with c1:
                    if st.button("Procesar", key=f"proc_{fname}", use_container_width=True):
                        try:
                            new_rows = PARSERS[ag_m]().parse(Path(data["path"]))
                            st.session_state["previews"][fname] = {
                                "agency": ag_m, "rows": new_rows,
                                "path": data["path"], "error": None,
                            }
                            st.rerun()
                        except Exception as e:
                            st.error(str(e))
                with c2:
                    if st.button("🗑️ Descartar", key=f"del_nd_{fname}", use_container_width=True):
                        del st.session_state["previews"][fname]
                        st.rerun()
                continue

            # ── Verificación IA (PDFs) ────────────────────────────────────────
            if is_pdf:
                if ai_key not in st.session_state:
                    # Verificar automáticamente al cargar
                    _run_ai_verification(Path(data["path"]), rows, agency, ai_key)
                    st.rerun()
                else:
                    _show_ai_result(st.session_state[ai_key], rows)

            # ── Tabla de vista previa ─────────────────────────────────────────
            df_rows = pd.DataFrame(rows)
            cols_show = [c for c in [
                "expedicion_agencia", "albaran_doccia", "pedido_doccia",
                "fecha_envio", "destinatario", "destino",
                "bultos", "kilos",
                "portes", "combustible", "seguro", "otros", "total_facturado",
                "estado_cruce",
            ] if c in df_rows.columns]
            rename_cols = {
                "expedicion_agencia": "Expedición", "albaran_doccia": "Albarán Doccia",
                "pedido_doccia": "Pedido", "fecha_envio": "Fecha", "destinatario": "Destinatario",
                "destino": "Destino", "bultos": "Bultos", "kilos": "Kg",
                "portes": "Portes €", "combustible": "RC Combustible €",
                "seguro": "Seguro €", "otros": "Contrareembolso €",
                "total_facturado": "Total €", "estado_cruce": "Estado",
            }
            df_show = (df_rows[cols_show] if cols_show else df_rows).rename(columns=rename_cols)
            st.dataframe(
                df_show,
                use_container_width=True, hide_index=True,
                height=min(400, 55 + len(rows) * 35),
            )

            # ── Botones acción ────────────────────────────────────────────────
            c1, c2, c3, _ = st.columns([1.2, 1, 1, 2])
            with c1:
                if st.button(f"💾 Guardar {len(rows):,} líneas", key=f"save_{fname}",
                             type="primary", use_container_width=True):
                    db = SessionLocal()
                    try:
                        save_bar = st.progress(0, text="Guardando…")
                        # ── Deduplicación: si esta factura ya estaba guardada,
                        #    borramos sus líneas antiguas y las reemplazamos ──────
                        factura_num = next((r.get("factura") for r in rows if r.get("factura")), None)
                        n_previas = 0
                        if factura_num:
                            n_previas = db.query(Albaran).filter(
                                Albaran.agencia == agency,
                                Albaran.factura == factura_num,
                            ).count()
                            if n_previas:
                                db.query(Albaran).filter(
                                    Albaran.agencia == agency,
                                    Albaran.factura == factura_num,
                                ).delete()
                        for j, row in enumerate(rows):
                            save_bar.progress((j + 1) / len(rows))
                            db.add(Albaran(
                                archivo_origen=fname,
                                agencia_detectada=agency,
                                **{k: row.get(k) for k in [
                                    "agencia", "factura", "fecha_envio", "expedicion_agencia",
                                    "albaran_doccia", "pedido_doccia", "destino", "destinatario",
                                    "bultos", "kilos", "peso_facturable", "portes", "combustible",
                                    "seguro", "reexpedicion", "otros", "total_facturado",
                                    "estado_cruce", "observaciones",
                                ]}
                            ))
                        db.commit()
                        save_bar.empty()
                        del st.session_state["previews"][fname]
                        st.session_state.pop(ai_key, None)
                        dup_msg = (f" (reemplazadas {n_previas:,} líneas previas de la factura {factura_num})"
                                   if n_previas else "")
                        st.session_state["carga_guardada"] = (
                            f"✅ {fname} — {len(rows):,} líneas guardadas{dup_msg}."
                        )
                        st.session_state["carga_upload_key"] += 1
                        st.rerun()
                    except Exception as e:
                        db.rollback()
                        st.error(f"Error al guardar: {e}")
                    finally:
                        db.close()
            with c2:
                if is_pdf and st.button("🔄 Re-verificar IA", key=f"reverify_{fname}",
                                        use_container_width=True):
                    st.session_state.pop(ai_key, None)
                    st.rerun()
            with c3:
                if st.button("🗑️ Descartar", key=f"del_{fname}", use_container_width=True):
                    del st.session_state["previews"][fname]
                    st.session_state.pop(ai_key, None)
                    st.session_state["carga_upload_key"] += 1
                    st.rerun()
